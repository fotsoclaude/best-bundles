from openpyxl import load_workbook
from pathlib import Path
import logging
import json
import shutil

BASE_DIR = Path(__file__).resolve().parent.parent
STORAGE_DIR = BASE_DIR / "storage"


def set_logging():
    logging.basicConfig(level=logging.INFO,
                        filename=BASE_DIR / "api.log",
                        filemode="w",
                        format='%(asctime)s - %(levelname)s - %(message)s')


def dict_sort_by_value(datas, reverse=False):
    return dict(sorted(datas.items(), key=lambda item: item[1], reverse=reverse))


def set_max_val_prio_to_one(datas):
    maximum = 0
    for val in datas.values():
        maximum = val
        break

    prios = []
    for k, v in datas.items():
        val = v / maximum
        datas[k] = val
        prios.append(val)
    return prios


def load_file(excel_file=BASE_DIR / 'helpers' / 'test.xlsx'):
    """Load bundles store in an Excel file into json with the name of each sheet.
    Each bundle's operator is store in a sheet with the heading:
     ---------------------------------------------
    | Name | Sms | Call | Data | Validity | Amount |
     ---------------------------------------------

    Args: excel_file (str, optional): Excel file containing differents bundles' operator per sheet.
    Defaults to 'test.xlsx'.
    """

    # Set storage
    shutil.rmtree(STORAGE_DIR, ignore_errors=True)
    STORAGE_DIR.mkdir(exist_ok=True)

    # Set logging
    set_logging()

    # load the spreadsheet
    wb = load_workbook(excel_file)

    for sheetname in wb.sheetnames:

        sheet = wb[sheetname]

        # Operator bundle empty file creation
        bundle_file = STORAGE_DIR / f"{sheetname.lower().replace(' ', '_')}.json"
        bundle_file.touch()
        bundle_file.write_text("{}")
        logging.info("Empty bundle file created.")

        row_count = sheet.max_row
        column_count = sheet.max_column

        headers = ["name", "sms", "call", "data", "validity", "amount"]
        with open(bundle_file, 'r') as f:
            bundles = json.load(f)

        for i in range(2, row_count + 1):
            bundle = {}
            maximum = 0
            for j in range(1, column_count + 1):
                data = sheet.cell(row=i, column=j).value

                # reformat data cells from str to int according to Mo or Go
                if j == 4:
                    try:
                        data_temp = data.split(" ")
                        if data_temp[1]:
                            bundle["data"] = int(float(data_temp[0].replace(",", ".")) * 1000)
                    except AttributeError:
                        bundle["data"] = data

                # reformat validity from str to int removing "jours"
                if j == 5:
                    try:
                        data_temp = data.split(" ")
                        if data_temp[1]:
                            data = int(data_temp[0])
                    except IndexError:
                        logging.debug("No 'jours' string on this validity.")

                # Add line's values in a dict
                if j != 4:
                    bundle[headers[j - 1]] = data

            # Add bundle info in the dict of bundles
            key = bundle.pop('name')
            bundles[key] = bundle.copy()

            # Sort the bundle items and assign them priorities
            del bundle['amount'], bundle['validity']
            sorted_bundle_items = dict_sort_by_value(bundle, reverse=True)
            priorities = set_max_val_prio_to_one(sorted_bundle_items)

            # Definition of an item priority in the bundle
            if priorities[0] == priorities[1]:
                if priorities[2] != 0 and priorities[2] != 1:
                    priorities[2] = 2
            elif priorities[1] == priorities[2] and priorities[1] != 0:
                priorities[1] = priorities[2] = 2
            elif priorities[1] == priorities[2] and priorities[1] == 0:
                pass
            elif priorities[2] == 0:
                priorities[1] = 2
            else:
                priorities[1] = 2
                priorities[2] = 3

            for k in range(0, 3):
                priorities[k] = int(priorities[k])

            priority_counter = 0
            for k, v in sorted_bundle_items.items():
                sorted_bundle_items[k] = priorities[priority_counter]
                priority_counter += 1

            bundles[key]["priorities"] = sorted_bundle_items

            with open(bundle_file, "w", encoding="utf-8") as f:
                json.dump(bundles, f, indent=4, ensure_ascii=False)
            logging.info(f"Bundle {key} of {sheetname} saved.")


def bundles_in_operator(operator_file, amount=1000000000, validity=1):
    available_bundles = {}
    with open(STORAGE_DIR / operator_file, "r") as f:
        bundles = json.load(f)
        for name, bundle in bundles.items():
            if (bundle['amount'] <= amount) and (bundle['validity'] >= validity):
                available_bundles[name] = bundle
    return available_bundles


def set_user_priorities(datas):
    datas = dict_sort_by_value(datas)
    maximum = 1
    while 0 in datas.values():
        for k, v in datas.items():
            del datas[k]
            break

    # diviser toutes les priorités par la priorité max
    for k, v in datas.items():
        maximum = v

    for k, v in datas.items():
        datas[k] = int(v / maximum)
        break
    logging.debug(f"New prios - {datas}")
    return datas


# Retourne la liste des meilleurs forfaits pour un opérateur
def best_bundle_in_operator(amount, sms=0, call=0, data=0, validity=1, cache=None, file=''):
    if cache is None:
        cache = {}
    logging.debug(f"Cache - {cache}")
    if sms == 0 and call == 0 and data == 0:
        return cache
    else:
        # Set user's priorities
        user_priorities = {'sms': sms, 'call': call, 'data': data}
        user_priorities = set_user_priorities(user_priorities)

        # Get the available bundles to the user's amount
        bundles = bundles_in_operator(file, amount, validity)

        # Filtrer la liste des forfaits pour garder ceux ayant la même priorité 1 que les choix de l'utilisateur
        can_do_bundles = tops = top = {}
        for k, v in user_priorities.items():
            for name, bundle in bundles.items():
                logging.debug(f"Available - '{name}': {bundle}")
                if bundle['priorities'][k] == 1:
                    can_do_bundles[name] = bundle
                    tops[name] = bundle[k] / bundle['priorities'][k]
                    break
            break

        tops = dict(sorted(tops.items(), key=lambda item: item[1], reverse=True))
        logging.debug(f"Tops - {can_do_bundles}")
        # Si on a une priorité dans la liste des priorités on ajoute le forfait et on retourne l'ensemble des forfaits
        # Si non, on cherche d'autres forfaits maximisant le reste des priorités
        top_name = ""
        for name, bundle in tops.items():
            top_name = name
            top[name] = bundle
            cache[top_name] = bundles[top_name]
            break
        # S'il veut maximiser un seul aspet du forfait, on lui le retourne LE forfait qui maximise son argent
        if len(user_priorities.keys()) == 1:
            logging.info(f"Le(s) meilleur(s) forfait(s): {cache}")
            return cache
        else:
            # On récupère les prioritées restantes.
            new_user_priorities = user_priorities.copy()
            top_bundle = bundles[top_name]
            bundle_priorities = top_bundle['priorities']
            prio_to_del = []
            for prio, prio_val in new_user_priorities.items():
                if prio_val == 1:
                    prio_to_del.append(prio)

            for prio in prio_to_del:
                del new_user_priorities[prio]
            logging.debug(f"Remaining prios - {new_user_priorities}")

            for prio in new_user_priorities.values():
                amount = amount - bundles[top_name]['amount']
                logging.info(f"Amount remaining - {amount}")
                return best_bundle_in_operator(amount=amount, **new_user_priorities, file=file,
                                               cache=cache, validity=validity)


def best_bundle_in_mango(amount, sms=0, call=0, data=0, validity=1, cache=None):
    set_logging()
    return best_bundle_in_operator(amount, sms, call, data, validity, cache, file='mango_forfaits.json')


def best_bundle_in_hemle(amount, sms=0, call=0, data=0, validity=1, cache=None):
    set_logging()
    return best_bundle_in_operator(amount, sms, call, data, validity, cache, file='hemle_forfaits.json')


if __name__ == "__main__":
    # load_file()
    mango = best_bundle_in_mango(amount=1000, sms=1, data=2, call=3, validity=7)
    hemle = best_bundle_in_hemle(amount=1000, sms=3, data=3, call=3, validity=7)
    for name, bundle in mango.items():
        print(
            f"Nom: {name} - Prix: {bundle['amount']} - Datas: {bundle['data']} - SMS: {bundle['sms']} - Call: {bundle['call']}")
    print("----" * 20)
    for name, bundle in hemle.items():
        print(
            f"Nom: {name} - Prix: {bundle['amount']} - Datas: {bundle['data']} - SMS: {bundle['sms']} - Call: {bundle['call']}")
